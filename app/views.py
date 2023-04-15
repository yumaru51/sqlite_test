from django.shortcuts import render, HttpResponse
from django.apps import apps  # app情報を取得
from django.db.models import Model  # Model情報を取得
from quality_change_management.models import TargetMaster, PageMaster, ActionMaster,\
    StepMaster, StepPageEntryMaster, StepDisplayPage, StepChargeDepartment, StepRelation, StepAction,\
    Log, Request, Quality, Safety, Progress
from fms.models import User, DepartmentMaster, DivisionMaster, UserAttribute
from pathlib import Path
from datetime import date, timezone, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import os
from socket import gethostname
host_name = gethostname()


def get_excluded_app_list():

    excluded_apps = [
        'django.contrib.admin',
        'django.contrib.auth',
        'django.contrib.contenttypes',
        'django.contrib.sessions',
        'django.contrib.messages',
        'django.contrib.staticfiles']
    app_configs = [app_config for app_config in apps.get_app_configs() if app_config.name not in excluded_apps]
    return app_configs


def output_test(request):

    data = {
        'app_list': get_excluded_app_list(),
        'output': TargetMaster.objects.all,
        'message': 'データを出力しました'
    }
    return render(request, 'app/app.html', data)


def delete_test(request):

    # 外部参照のため、DELETEの順番に注意
    StepAction.objects.all().delete()
    StepRelation.objects.all().delete()
    StepChargeDepartment.objects.all().delete()
    StepDisplayPage.objects.all().delete()
    StepPageEntryMaster.objects.all().delete()
    StepMaster.objects.all().delete()
    ActionMaster.objects.all().delete()
    PageMaster.objects.all().delete()
    TargetMaster.objects.all().delete()

    UserAttribute.objects.all().delete()
    DivisionMaster.objects.all().delete()
    DepartmentMaster.objects.all().delete()
    User.objects.all().delete()

    data = {
        'app_list': get_excluded_app_list(),
        'message': 'データを削除しました'
    }
    return render(request, 'app/app.html', data)


def import_excel_to_model(model: Model, file_name, file):
    print(file_name + 'を処理しています')
    fields = model._meta.fields
    kwargs = {}

    # ファイルに保存済みのExcelファイル読込
    # workbook = openpyxl.load_workbook(output_folder_path + '/' + file_name + '.xlsx')
    # メモリ上に一時ファイルを保存し、Openpyxlで読み込む
    workbook = load_workbook(file, read_only=True)
    worksheet = workbook.active

    # 項目の取得
    column_names = []
    for cell in worksheet[1]:
        column_names.append(cell.value)

    # データの取得
    for column_data in worksheet.iter_rows(min_row=2):
        for i in range(len(column_names)):
            # print('項目名：' + column_names[i] + '値：' + column_data[i].value)

            # 外部参照キーの場合、参照元テーブルのインスタンスを格納
            if fields[i].get_internal_type() == 'ForeignKey':
                try:
                    # 参照元テーブルのインスタンス取得
                    if column_data[i].value is not None:
                        ref_pk = column_data[i].value
                        ref_model = fields[i].remote_field.model
                        ref_instance = ref_model.objects.get(pk=ref_pk)
                        kwargs[column_names[i]] = ref_instance
                    else:
                        kwargs[column_names[i]] = None
                except ValueError:
                    continue
            elif fields[i].get_internal_type() == 'DateTimeField':
                try:
                    # 日付データを文字型からフォーマット変換
                    if column_data[i].value is not None:
                        date_time_field = column_data[i].value
                        kwargs[column_names[i]] = date_time_field.astimezone(timezone(timedelta(hours=9), 'JST'))
                    else:
                        kwargs[column_names[i]] = None
                except ValueError:
                    continue
            else:
                kwargs[column_names[i]] = column_data[i].value

        # 1行ずつIMPORT処理
        # globals()[file_name].objects.create(**{column_names[i]: column_data[i].value})  # 辞書に変換する構文
        # kwargs = dict(zip(column_names, column_datas))
        # globals()[file_name].objects.create(**kwargs)  # これだとライブラリimportが必要。
        model.objects.create(**kwargs)

    return


def import_data(request):
    today = str(date.today())
    app_name = request.POST['app_name']

    # ファイルが1件もない
    if request.FILES.__len__() == 0:
        data = {
            'app_list': get_excluded_app_list(),
            'message': 'フォルダを選択していません。'
        }
        return render(request, 'app/app.html', data)

    # import用EXCELファイルの読み込み
    else:
        for file in request.FILES.getlist('import_file'):
            # ファイル名をファイル名部分と拡張子部分で分ける
            file_name, ext = os.path.splitext(file.name)

            # ModelClass指定でModel情報取得
            model_data = apps.get_model(app_name, file_name)  # "app_name.ModelName"
            import_excel_to_model(model_data, file_name, file)

    print('import完了')

    data = {
        'app_list': get_excluded_app_list(),
        'message': 'データをimportしました'
    }

    return render(request, 'app/app.html', data)


# 外部参照キー対応
def export_model_to_excel(model, output_file_path):
    # get all fields
    fields = model._meta.fields

    # create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sheet1'

    # create custom named style 別に書式とかはいらないけど、、、
    date_style = NamedStyle(name='date_style')
    date_style.font = Font(name='Calibri', size=11, bold=False)
    date_style.alignment = Alignment(horizontal='left', vertical='center')
    date_style.number_format = 'yyyy/mm/dd'

    datetime_style = NamedStyle(name='datetime_style')
    datetime_style.font = Font(name='Calibri', size=11, bold=False)
    datetime_style.alignment = Alignment(horizontal='left', vertical='center')
    datetime_style.number_format = 'yyyy/m/d hh:mm:ss'

    # write headers
    for i, field in enumerate(fields):
        col_letter = get_column_letter(i+1)
        worksheet[col_letter + '1'] = field.name

    # write data
    row = 2
    for obj in model.objects.all():
        col = 1
        for field in fields:
            if field.get_internal_type() == 'ForeignKey':
                # check if foreign key value can be converted to
                try:
                    # 参照先テーブルの外部参照キー項目
                    foreignkey_field = getattr(obj, field.name)

                    # 参照元テーブルの主キー取得
                    if foreignkey_field is not None:
                        # ①参照元テーブルの主キー値取得
                        foreignkey_value = foreignkey_field.pk
                        # ②参照元テーブルの主キーを取得してから値を取得
                        # pk_name = foreignkey_field._meta.pk.name  # たぶん参照元テーブルとかも取れる？
                        # foreignkey_value = foreignkey_field.__getattribute__(pk_name)  # 動的にするやつ？
                    else:
                        foreignkey_value = ''
                except ValueError:
                    continue
                # write the foreign key value to the worksheet
                worksheet.cell(row=row, column=col, value=foreignkey_value)

            elif field.get_internal_type() == 'DateField':
                try:
                    date_value = getattr(obj, field.name)
                except ValueError:
                    continue
                # write datetime value with custom style
                cell = worksheet.cell(row=row, column=col, value=date_value)
                cell.style = date_style

            elif field.get_internal_type() == 'DateTimeField':
                try:
                    datetime_field = getattr(obj, field.name)
                    if datetime_field is not None:
                        datetime_value = datetime_field.replace(tzinfo=None)
                    else:
                        datetime_value = ''
                except ValueError:
                    continue
                # write datetime value with custom style
                cell = worksheet.cell(row=row, column=col, value=datetime_value)
                cell.style = datetime_style

            else:
                # write the field value to the worksheet
                worksheet.cell(row=row, column=col, value=getattr(obj, field.name))

            col += 1
        row += 1

    # save workbook to output file path
    workbook.save(output_file_path)


# Model名を指定して、Excelファイルに出力する
def export_data(request):
    today = str(date.today())
    app_name = request.POST['app_name']
    if host_name == 'I7781DN7':
        output_folder_path = 'C:/Users/y-kawauchi/python_tool_development/sqlite_test/' + today + '_' + app_name
    else:
        output_folder_path = '/Users/kawauchiyuuki/PycharmProjects/sqlite_test/' + today + '_' + app_name

    # フォルダを作成する
    Path(output_folder_path).mkdir(parents=True, exist_ok=True)

    # ModelClass指定でModel情報取得
    # model_data = apps.get_model("app_name.ModelName")  # "app_name.ModelName"
    # プロジェクトの全Model情報取得
    # model_list = apps.get_models()
    # アプリ毎のModel情報取得
    app = apps.get_app_config(app_name)  # アプリケーション名で指定したアプリケーションオブジェクトを取得
    model_list = app.get_models()  # アプリケーション内の全モデルを取得

    # モデル名のリストを取得、export
    for model in model_list:
        print('App名：' + str(app) + ', Model名：' + model.__name__)
        model_data = app.get_model(model.__name__)  # "app_name.ModelName"
        output_file_path = output_folder_path + '/' + model.__name__ + ".xlsx"
        export_model_to_excel(model_data, output_file_path)

    data = {
        'app_list': get_excluded_app_list(),
        'app_name': app_name,
        'message': app_name + 'をexportしました'
    }

    return render(request, 'app/app.html', data)


def test_function(request):

    kwargs = {
        'id': 1,
        'division': 'TIODIV',
        'department': 'S',
        'user': 'end',
        'change_target': '[method, facility, ]',
        'others': None,
        'title': '1B2BC工場\u3000H-N廃酸運用方法の変更',
        'outline': '①\t1次洗浄濾液をタイマー管理でH・N廃酸に振り分けていったが、H-N廃酸の切替をなくし全量N廃酸とする。\n②\tW-10タンクとW-04タンクを底配管で連通させて、加水分解タンクの希釈水タンクの容量を20→40m3としオーバーフローとしてN廃酸へロスしていたものを削減する。',
        'level': None,
        'treatment': None,
        'safety_aspect': '0',
        'quality_aspect': '0',
        'delivery_date': 42552,
        # 'delivery_date_start': None,
        'delivery_date_end': None,
        'level2': '継続',
        'others2': None,
        'completion_date': None,
        'application_date': None,
        'education_management_system_id': None
    }
    Request.objects.create(**kwargs)

    data = {
        'app_list': get_excluded_app_list(),
        'message': 'テスト完了'
    }

    return render(request, 'app/app.html', data)
